from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from .models import *
from django.http import HttpResponse

from .forms import UploadFileForm
import csv
from openpyxl import load_workbook
from django.core.paginator import Paginator

# Create your views here.
def index(request):
    """
    Display the home page of the web app
    """
    return render(request, 'index.html')

def register_user(request):
    """
    Create a new user
    """
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if User.objects.filter(username=username).exists():
            messages.error(request, "This username is already taken.")
            return redirect('register')

        if User.objects.filter(email=email).exists():
            messages.error(request, "This email is already taken.")
            return redirect('register')

        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect('register')

        try:
            user = User.objects.create_user(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                password=password1
            )
            user.save()
            messages.success(request, "You have successfully registered! Kindly login below!")
            return redirect('login')
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return render(request, 'register.html')
    else:
        return render(request, 'register.html')

def login_user(request):
    """
    Logs in a user
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, "You have been logged in successfully")
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password.")
            return redirect('login')
    return render(request, 'login.html')

def logout_user(request):
    """
    Logs out a user
    """
    logout(request)
    messages.success(request ,'You have been logged out successfully ')
    return redirect('login')

@login_required(login_url='login/')
def dashboard(request):
    """
    Display stats of the system
    """
    return render(request, 'dashboard.html')

@login_required(login_url='login/')
def get_users_book(request):
    """
    Returns all the book categories
    """
    user = request.user
    # books = Book.objects.all()
    books = Book.objects.filter(user=user)

    paginator = Paginator(books, 10)  # Show 10 books per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    return render(request, 'book-list.html', {'books': books, 'page_obj': page_obj})


    # Add pagination
    # return render(request, 'book-list.html', {'books': books})

@login_required(login_url='login/')
def book_details(request, pk):
    """
    Returns the details of a book
    """
    books = Book.objects.get(id=pk)
    return render(request, "book-detail.html", {'books': books})

@login_required(login_url='login/')
def add_book(request):
    """
    Adds a new book category
    """
    if request.method == 'POST':
        uid = request.POST.get('uid')
        title = request.POST.get('title')
        subtitle = request.POST.get('subtitle')
        author = request.POST.get('author')
        publisher = request.POST.get('publisher')
        published_date = request.POST.get('published_date')
        category = request.POST.get('category')
        distribution_expense = request.POST.get('distribution_expense')

        try:
            book = Book.objects.create(
                uid=uid,
                title=title,
                subtitle=subtitle,
                author=author,
                publisher=publisher,
                published_date=published_date,
                category=category,
                distribution_expense=distribution_expense
                )
            messages.success(request, "You have added a new record successfully")
            
            return redirect('books')
        except Exception as e:
            # return HttpResponse(f"Error: {e}")
            messages.error(request, f"Error: {e}")

    return render(request, 'add-book.html')

@login_required(login_url='login/')
def update_book(request, pk):
    """
    Updates a book category
    """
    book = Book.objects.get(id=pk)
    if request.method == 'POST':
        uid = request.POST.get('uid')
        title = request.POST.get('title')
        subtitle = request.POST.get('subtitle')
        author = request.POST.get('author')
        publisher = request.POST.get('publisher')
        published_date = request.POST.get('published_date')
        category = request.POST.get('category')
        distribution_expense = request.POST.get('distribution_expense')
        try:
            book.uid = uid
            book.title = title
            book.subtitle = subtitle
            book.author = author
            book.publisher = publisher
            book.published_date = published_date
            book.category = category
            book.distribution_expense = distribution_expense
            book.save()
            messages.success(request, "You have updated a record successfully")
            return redirect('book-detail', pk=book.pk)
        
        except Exception as e:
            return HttpResponse(f"Error: {e}")
            # messages.error(request, f"Error: {e}")

    return render(request, 'update-book.html', {'book': book})

@login_required(login_url='login/')
def delete_book(request, pk):
    """
    Deletes a book category
    """
    delete_book = Book.objects.get(id=pk)
    try:
        delete_book.delete()
        messages.success(request, "You have deleted a record successfully")
        return redirect('books')
    except Exception as e:
        return HttpResponse(f"Error: {e}")

def handle_uploaded_file(f, user):
    # Handle file processing based on the file type
    if f.name.endswith('.csv'):
        # Read CSV file
        file_content = f.read().decode('utf-8').splitlines()
        reader = csv.DictReader(file_content)
        
        for row in reader:
            Book.objects.create(
                uid=row['id'],
                title=row['title'],
                subtitle=row['subtitle'],
                author=row['author'],
                publisher=row['publisher'],
                published_date=row['published_date'],
                category=row['category'],
                distribution_expense=row['distribution_expense'],
                user=user,  # Associate the book with the user
            )
    elif f.name.endswith('.xlsx'):
        # Read Excel file (.xlsx)
        wb = load_workbook(f)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skips the header row
            Book.objects.create(
                uid=row[0],
                title=row[1],
                subtitle=row[2],
                author=row[3],
                publisher=row[4],
                published_date=row[5],
                category=row[6],
                distribution_expense=row[7],
                user=user,  # Associate the book with the user
            )
    else:
        raise ValueError('Invalid file format')

@login_required(login_url='login/')
def upload_file(request):
    """
    Uploads a file
    """
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            # Process the uploaded file
            handle_uploaded_file(request.FILES['file'], request.user)  # Pass the user
            
            # Save the uploaded file instance
            UploadedFile.objects.create(file=request.FILES['file'], user=request.user)
            messages.success(request, "File uploaded and data saved successfully")
            return redirect('books')
            # return HttpResponse('File uploaded and data saved successfully.')
        except Exception as e:
            return HttpResponse(f'An error occurred: {str(e)}')
    else:
        form = UploadFileForm()  # Initialize the form for the file upload
    return render(request, 'upload.html', {'form': form})
